#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Вывод результатов указанных спортсменов за все турниры в базе.

Список ФИО задаётся в переменной ATHLETE_NAMES ниже или в файле (--file).
Совпадение по нормализованному ФИО (порядок слов не важен).

Использование:
  python scripts/athlete_results_report.py
  python scripts/athlete_results_report.py --file path/to/names.txt
  python scripts/athlete_results_report.py --csv -o report.csv
  python scripts/athlete_results_report.py --excel -o report.xlsx

Запуск из корня проекта.
"""

import csv
import os
import re
import sys

project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from app_factory import create_app
from extensions import db
from models import Athlete, Event, Category, Participant

# Список ФИО спортсменов для отчёта (уникальные по нормализованному ключу)
ATHLETE_NAMES = [
    "Бурданова Александра Владимировна",
    "Заикина Алиса Степановна",
    "Ильиных Анна Максимовна",
    "Шарипова София Айдаровна",
    "Белова Елизавета Антоновна",
    "Вачасова Ева Владимировна",
    "Значкова Полина Артемовна",
    "Капичникова Мария Евгеньевна",
    "Мурашова Злата Вячеславовна",
    "Плесская Мария Юрьевна",
    "Аксенова Мария Андреевна",
    "Аскарова Илина Ильшатовна",
    "Базылюк Маргарита Андреевна",
    "Барышева Анастасия Алексеевна",
    "Бойцова Арина Станиславовна",
    "Брановицкая Арина Максимовна",
    "Буря Злата Олеговна",
    "Галушкина Мария Сергеевна",
    "Герцог Арина Антоновна",
    "Голубева Елизавета Андреевна",
    "Гончарова Аврора Егоровна",
    "Гречихина София Алексеевна",
    "Дзепка София Александровна",
    "Докукина Анна Андреевна",
    "Дронова Алёна Дмитриевна",
    "Дьякова Таисия Петровна",
    "Иванова Яна Михайловна",
    "Кварацхелия Ева Георгиевна",
    "Корнаухова Алиса Ильинична",
    "Коротких Николина Юрьевна",
    "Корчажникова Екатерина Алексеевна",
    "Корякова Майя Дмитриевна",
    "Кравчина Варвара Евгеньевна",
    "Кудряшова Елизавета Евгеньевна",
    "Курдюмова Дария Алексеевна",
    "Лабуткина Софья Алексеевна",
    "Лебедева Дарья Евгеньевна",
    "Литвинчёва Валерия Денисовна",
    "Лукашова Валерия Дмитриевна",
    "Мартыненко Полина Евгеньевна",
    "Маскайкина Ника Анатольевна",
    "Мильто Алиса Вадимовна",
    "Мильто Диана Вадимовна",
    "Митрофанова Ариадна Андреевна",
    "Михайлова Виктория Сергеевна",
    "Мураткалиева Дарья Сергеевна",
    "Мышкина Елизавета Алексеевна",
    "Назипова Камила Ильясовна",
    "Невзорова Варвара Вячеславовна",
    "Нестеренок Елизавета Дмитриевна",
    "Николаева Алина Павловна",
    "Павликова Полина Владимировна",
    "Погребицкая Таисия Сергеевна",
    "Полуэктова Варвара Ивановна",
    "Поступаева Дарья Юрьевна",
    "Принева Алена Максимовна",
    "Радионова Юлия Ивановна",
    "Сарафанова Майя Андреевна",
    "Сарновская София Валерьевна",
    "Сарычева Марина Викторовна",
    "Скучалина София Дмитриевна",
    "Смагина Софья Сергеевна",
    "Степанова Александра Михайловна",
    "Стоцкая Екатерина Евгеньевна",
    "Стрельцова Виктория Вадимовна",
    "Суханова Лидия Алексеевна",
    "Тарасенко Ева Михайловна",
    "Телятникова Алёна Алексеевна",
    "Титова Екатерина Дмитриевна",
    "Токарева Александра Дмитриевна",
    "Третьякова София Игоревна",
    "Трифонова Маргарита Владимировна",
    "Уткина Анастасия Игоревна",
    "Федоркова Софья Евгеньевна",
    "Фирсунина Глория Игоревна",
    "Чиннова Анастасия Андреевна",
    "Чистякова Алиса Андреевна",
    "Шарошкина Анастасия Васильевна",
    "Шиляева Ульяна Сергеевна",
]


def normalize_fio_key(s):
    """Нормализация ФИО для сравнения: нижний регистр, один пробел, ключ = кортеж слов."""
    if not s or not isinstance(s, str):
        return None
    s = re.sub(r'\s+', ' ', s.strip()).lower()
    if not s:
        return None
    return tuple(sorted(s.split()))


def build_athlete_index(app):
    """Индекс: normalized_key -> [(athlete_id, full_name), ...]"""
    index = {}
    with app.app_context():
        for a in Athlete.query.all():
            name = a.full_name
            key = normalize_fio_key(name)
            if key:
                index.setdefault(key, []).append((a.id, name))
    return index


def get_participations_for_athlete(app, athlete_id):
    """Все участия спортсмена: (Event, Category, Participant), по убыванию даты."""
    with app.app_context():
        return db.session.query(Event, Category, Participant).join(
            Category, Participant.category_id == Category.id
        ).join(
            Event, Category.event_id == Event.id
        ).filter(
            Participant.athlete_id == athlete_id
        ).order_by(Event.begin_date.desc()).all()


def format_date_range(event):
    """Строка дат турнира."""
    if not event.begin_date:
        return ""
    start = event.begin_date.strftime('%d.%m.%Y')
    if event.end_date and event.end_date != event.begin_date:
        return f"{start} — {event.end_date.strftime('%d.%m.%Y')}"
    return start


def format_place(participant):
    """Место или статус (WD/R и т.д.)."""
    if participant.total_place is not None:
        return str(participant.total_place)
    if participant.status == 'W':
        return "WD"
    if participant.status == 'R':
        return "R"
    return "—"


def format_points(participant):
    """Очки или прочерк."""
    if participant.total_points is not None:
        return f"{participant.total_points:.2f}"
    return "—"


def build_excel(results):
    """
    Строит таблицу для Excel: строки — спортсмены (ФИО), столбцы — турниры,
    в ячейках — баллы (если несколько участий в одном турнире — через запятую).
    Возвращает (headers, rows).
    """
    from datetime import date
    # Собираем все турниры: (event_id, event_name, begin_date), сортируем по дате
    events_seen = {}
    for r in results:
        for event, category, participant in r['participations']:
            if event.id not in events_seen:
                events_seen[event.id] = (event.name or f'Турнир {event.id}', event.begin_date)
    events_sorted = sorted(
        events_seen.items(),
        key=lambda x: (x[1][1] or date(1900, 1, 1), x[0])
    )
    event_columns = [(eid, ename) for eid, (ename, _) in events_sorted]

    headers = ['ФИО'] + [ename for _, ename in event_columns]
    event_ids = [eid for eid, _ in event_columns]
    rows = []
    for r in results:
        points_by_event = {}
        for event, category, participant in r['participations']:
            eid = event.id
            pts = f"{participant.total_points:.2f}" if participant.total_points is not None else "—"
            points_by_event.setdefault(eid, []).append(pts)
        row = [r['full_name']]
        for eid in event_ids:
            vals = points_by_event.get(eid, [])
            row.append(", ".join(vals) if vals else "")
        rows.append(row)
    return headers, rows


def write_excel_file(path, headers, rows):
    """Пишет в .xlsx: первая строка — заголовки (ФИО, названия турниров), далее — данные."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    ws.column_dimensions['A'].width = 35
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20
    wb.save(path)


def load_names_from_file(path):
    """Читает ФИО из файла: по одному на строку, пустые и заголовок 'ФИО' пропускаются."""
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line == 'ФИО':
                continue
            yield line


def main():
    import argparse
    parser = argparse.ArgumentParser(description='Отчёт по результатам спортсменов за все турниры')
    parser.add_argument('--file', '-f', help='Файл со списком ФИО (по одному на строку)')
    parser.add_argument('--csv', action='store_true', help='Вывести CSV в stdout или в файл -o')
    parser.add_argument('--excel', '-x', action='store_true', help='Выгрузить в Excel (.xlsx); файл задаётся через -o')
    parser.add_argument('-o', '--output', help='Файл для вывода (при --csv или --excel)')
    args = parser.parse_args()

    if args.file:
        if not os.path.isfile(args.file):
            print(f"Файл не найден: {args.file}", file=sys.stderr)
            sys.exit(1)
        requested = list(load_names_from_file(args.file))
    else:
        requested = list(ATHLETE_NAMES)

    # Уникальные ключи (сохраняем первое вхождение ФИО для отображения)
    seen_keys = set()
    unique_names = []
    for name in requested:
        key = normalize_fio_key(name)
        if key and key not in seen_keys:
            seen_keys.add(key)
            unique_names.append(name)

    app = create_app()
    index = build_athlete_index(app)

    # Собираем данные: (запрошенное_ФИО, [(athlete_id, full_name)], participations)
    results = []
    not_found = []

    for display_name in unique_names:
        key = normalize_fio_key(display_name)
        if not key:
            continue
        matches = index.get(key, [])
        if not matches:
            not_found.append(display_name)
            continue
        for athlete_id, full_name in matches:
            participations = get_participations_for_athlete(app, athlete_id)
            results.append({
                'requested_name': display_name,
                'athlete_id': athlete_id,
                'full_name': full_name,
                'participations': participations,
            })

    # Вывод
    if args.excel:
        out_path = args.output or 'athlete_results.xlsx'
        headers, rows = build_excel(results)
        write_excel_file(out_path, headers, rows)
        print(f"Excel записан: {out_path}", file=sys.stderr)
        if not_found:
            print(f"Не найдены в БД ({len(not_found)}): {', '.join(not_found)}", file=sys.stderr)
        return

    if args.csv:
        out = open(args.output, 'w', newline='', encoding='utf-8') if args.output else sys.stdout
        writer = csv.writer(out, delimiter=';')
        writer.writerow([
            'ФИО (запрос)', 'ФИО в БД', 'ID спортсмена',
            'Турнир', 'Дата', 'Категория', 'Место', 'Очки', 'Бесплатное участие', 'Статус'
        ])
        for r in results:
            for event, category, participant in r['participations']:
                writer.writerow([
                    r['requested_name'],
                    r['full_name'],
                    r['athlete_id'],
                    event.name or '',
                    format_date_range(event),
                    category.name if category else '',
                    format_place(participant),
                    format_points(participant),
                    'да' if participant.pct_ppname == 'БЕСП' else '',
                    participant.status or '',
                ])
        if args.output:
            out.close()
            print(f"CSV записан в {args.output}", file=sys.stderr)
        return

    # Текстовый вывод
    print("=" * 100)
    print("РЕЗУЛЬТАТЫ СПОРТСМЕНОВ ЗА ВСЕ ТУРНИРЫ В БАЗЕ")
    print("=" * 100)
    print(f"Запрошено уникальных ФИО: {len(unique_names)}")
    print(f"Найдено в БД: {len(results)} (не найдено: {len(not_found)})")
    print("=" * 100)

    for r in results:
        print(f"\n{'─' * 100}")
        print(f"  {r['full_name']}  (ID: {r['athlete_id']})")
        print(f"{'─' * 100}")
        if not r['participations']:
            print("  Участий в турнирах не найдено.")
            continue
        for event, category, participant in r['participations']:
            date_str = format_date_range(event)
            place_str = format_place(participant)
            points_str = format_points(participant)
            free = " [БЕСП]" if participant.pct_ppname == 'БЕСП' else ""
            print(f"  • {event.name or '—'}")
            print(f"      Дата: {date_str}  |  Категория: {category.name if category else '—'}  |  Место: {place_str}  |  Очки: {points_str}{free}")

    if not_found:
        print(f"\n{'=' * 100}")
        print("НЕ НАЙДЕНЫ В БАЗЕ:")
        for name in not_found:
            print(f"  {name}")
        print("=" * 100)


if __name__ == '__main__':
    main()
